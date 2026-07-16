"""
Excel export for invoices, quotes, and expenses using openpyxl.

Deep module: small interface (3 public methods), concentrated
spreadsheet formatting behind it.
"""

import io
import logging
from datetime import date, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Shared styles
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_HEADER_FILL = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BODY_FONT = Font(name="Calibri", size=10)
_MONEY_FORMAT = "#,##0.00"
_DATE_FORMAT = "DD MMM YYYY"
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="E5E7EB"),
)


class ExcelExporter:
    """Generate branded Excel workbooks for document exports."""

    # public interface

    def export_invoices(
        self,
        invoices: list,
        include_line_items: bool = False,
    ) -> bytes:
        """Export invoice records to .xlsx bytes."""
        headers = [
            "Invoice #",
            "Reference",
            "Customer",
            "Date",
            "Due Date",
            "Status",
            "Currency",
            "Subtotal",
            "Tax",
            "Total Due",
            "Amount Paid",
            "Balance Due",
        ]

        def row_fn(inv):
            return [
                inv.invoice_number,
                inv.invoice_reference,
                getattr(inv.customer, "display_name", str(inv.customer_id)),
                inv.transaction_date,
                inv.due_date,
                inv.status,
                inv.currency,
                inv.subtotal,
                inv.tax_total,
                inv.total_due,
                inv.amount_paid,
                inv.balance_due,
            ]

        money_cols = [8, 9, 10, 11, 12]  # 1-indexed: Subtotal..Balance
        date_cols = [4, 5]

        wb = self._build_workbook(
            sheet_name="Invoices",
            headers=headers,
            records=invoices,
            row_fn=row_fn,
            money_cols=money_cols,
            date_cols=date_cols,
        )

        if include_line_items and invoices:
            self._add_line_items_sheet(wb, invoices, "invoice")

        return self._to_bytes(wb)

    def export_quotes(
        self,
        quotes: list,
        include_line_items: bool = False,
    ) -> bytes:
        """Export quote records to .xlsx bytes."""
        headers = [
            "Quote #",
            "Reference",
            "Customer",
            "Date",
            "Due Date",
            "Status",
            "Currency",
            "Subtotal",
            "Tax",
            "Total Due",
        ]

        def row_fn(q):
            return [
                q.quote_number,
                q.quote_reference,
                getattr(q.customer, "display_name", str(q.customer_id)),
                q.transaction_date,
                q.due_date,
                q.status,
                q.currency,
                q.subtotal,
                q.tax_total,
                q.total_due,
            ]

        wb = self._build_workbook(
            sheet_name="Quotes",
            headers=headers,
            records=quotes,
            row_fn=row_fn,
            money_cols=[8, 9, 10],
            date_cols=[4, 5],
        )

        if include_line_items and quotes:
            self._add_line_items_sheet(wb, quotes, "quote")

        return self._to_bytes(wb)

    def export_expenses(
        self,
        expenses: list,
        include_line_items: bool = False,
    ) -> bytes:
        """Export expense records to .xlsx bytes."""
        headers = [
            "Expense #",
            "Reference",
            "Vendor",
            "Date",
            "Due Date",
            "Status",
            "Currency",
            "Subtotal",
            "Tax",
            "Total Due",
            "Amount Paid",
            "Balance Due",
        ]

        def row_fn(exp):
            vendor_name = getattr(exp, "vendor_name", None)
            if not vendor_name and hasattr(exp, "vendor"):
                vendor_name = getattr(exp.vendor, "vendor_name", str(exp.vendor_id))
            return [
                exp.expense_number,
                exp.expense_reference,
                vendor_name or str(exp.vendor_id),
                exp.expense_date,
                exp.due_date,
                exp.status,
                exp.currency,
                exp.subtotal,
                exp.tax_total,
                exp.total_due,
                exp.amount_paid,
                exp.balance_due,
            ]

        wb = self._build_workbook(
            sheet_name="Expenses",
            headers=headers,
            records=expenses,
            row_fn=row_fn,
            money_cols=[8, 9, 10, 11, 12],
            date_cols=[4, 5],
        )

        if include_line_items and expenses:
            self._add_line_items_sheet(wb, expenses, "expense")

        return self._to_bytes(wb)

    def export_purchase_orders(
        self,
        purchase_orders: list,
        include_line_items: bool = False,
    ) -> bytes:
        """Export purchase-order records to .xlsx bytes.

        Mirrors export_expenses (vendor-facing). No discount in v1, so the
        money columns are Subtotal / Tax / Total only.
        """
        headers = [
            "PO #",
            "Reference",
            "Vendor",
            "Order Date",
            "Delivery Date",
            "Status",
            "Currency",
            "Subtotal",
            "Tax",
            "Total",
        ]

        def row_fn(po):
            vendor_name = getattr(po, "vendor_name", None)
            if not vendor_name and hasattr(po, "vendor"):
                vendor_name = getattr(po.vendor, "vendor_name", str(po.vendor_id))
            return [
                po.po_number,
                po.po_reference,
                vendor_name or str(po.vendor_id),
                po.order_date,
                po.delivery_date,
                po.status,
                po.currency,
                po.subtotal,
                po.tax_total,
                po.total,
            ]

        wb = self._build_workbook(
            sheet_name="Purchase Orders",
            headers=headers,
            records=purchase_orders,
            row_fn=row_fn,
            money_cols=[8, 9, 10],
            date_cols=[4, 5],
        )

        if include_line_items and purchase_orders:
            # doc_type "po" → the sheet resolves the document number via
            # getattr(record, "po_number"), which is the PO's actual attr.
            self._add_line_items_sheet(wb, purchase_orders, "po")

        return self._to_bytes(wb)

    def export_purchase_order_payments(
        self,
        purchase_order,
        payments: list,
    ) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Payments"

        po_reference = getattr(purchase_order, "po_reference", "")
        vendor_name = getattr(purchase_order, "vendor_name", None)
        if not vendor_name and hasattr(purchase_order, "vendor"):
            vendor_name = getattr(
                purchase_order.vendor, "vendor_name", str(purchase_order.vendor_id)
            )

        title_text = f"{vendor_name}_{po_reference}" if vendor_name else po_reference

        # Determine currency formatting (fallback to standard $ if unknown, though frontend formats as needed)
        # Using a simple formatting string to show symbol + amount.
        currency = getattr(purchase_order, "currency", "USD")
        curr_map = {"USD": "$", "KES": "KES", "EUR": "€", "GBP": "£"}
        sym = curr_map.get(currency, currency)
        money_fmt = f'"{sym}" #,##0.00'

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Row 1: Title
        ws.merge_cells("A1:F1")
        title_cell = ws.cell(row=1, column=1, value=title_text)
        title_cell.font = Font(name="Calibri", bold=True, size=16, color="2C1A54")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row 2: Headers
        headers = [
            "Date",
            "Reference",
            "Amount",
            "Payment Reference",
            "Amount",
            "Notes",
        ]
        header_font = Font(name="Calibri", bold=True, color="912B90")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.border = thin_border

        row_idx = 3

        # Row 3: PO Row
        po_date = purchase_order.order_date
        po_total = purchase_order.total or Decimal("0.00")

        c_po_date = ws.cell(row=row_idx, column=1, value=po_date)
        c_po_date.number_format = _DATE_FORMAT
        ws.cell(row=row_idx, column=2, value=po_reference)
        c_po_amt = ws.cell(row=row_idx, column=3, value=po_total)
        c_po_amt.number_format = money_fmt

        for col_i in range(1, 7):
            ws.cell(row=row_idx, column=col_i).border = thin_border

        row_idx += 1

        # Payment Rows
        total_payments = Decimal("0.00")
        for p in payments:
            c_p_date = ws.cell(row=row_idx, column=1, value=p.payment_date)
            c_p_date.number_format = _DATE_FORMAT
            ws.cell(row=row_idx, column=4, value=p.reference or "")
            c_p_amt = ws.cell(row=row_idx, column=5, value=p.amount)
            c_p_amt.number_format = money_fmt
            ws.cell(row=row_idx, column=6, value=p.notes or "")

            for col_i in range(1, 7):
                ws.cell(row=row_idx, column=col_i).border = thin_border

            total_payments += p.amount
            row_idx += 1

        # Add 2 empty spacer rows
        for _ in range(2):
            for col_i in range(1, 7):
                ws.cell(row=row_idx, column=col_i).border = thin_border
            row_idx += 1

        # Totals block
        # PO Amount row
        c_po_label = ws.cell(row=row_idx, column=1, value="PO Amount")
        c_po_label.font = Font(name="Calibri", size=10)
        c_po_val = ws.cell(row=row_idx, column=3, value=po_total)
        c_po_val.number_format = money_fmt
        for col_i in range(1, 7):
            ws.cell(row=row_idx, column=col_i).border = thin_border
        row_idx += 1

        # Total Payments row
        c_tp_label = ws.cell(row=row_idx, column=1, value="Total Payments")
        c_tp_label.font = Font(name="Calibri", size=10)
        c_tp_val = ws.cell(row=row_idx, column=5, value=total_payments)
        c_tp_val.number_format = money_fmt
        for col_i in range(1, 7):
            ws.cell(row=row_idx, column=col_i).border = thin_border
        row_idx += 1

        # Balance row
        balance = po_total - total_payments
        color = "EF4444" if balance < 0 else "000000"

        c_bal_lbl = ws.cell(row=row_idx, column=1, value="Bal")
        c_bal_lbl.font = Font(name="Calibri", size=10, color=color)
        c_bal_val = ws.cell(row=row_idx, column=5, value=balance)
        c_bal_val.number_format = money_fmt
        c_bal_val.font = Font(name="Calibri", size=10, color=color)

        for col_i in range(1, 7):
            ws.cell(row=row_idx, column=col_i).border = thin_border

        # Set column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 25

        return self._to_bytes(wb)

    def export_vendor_card(
        self,
        *,
        title: str,
        vendor_name: str,
        currency: str,
        items: list,
        totals: dict,
    ) -> bytes:
        """Export a vendor detail card (POs / Payments / Bills) to .xlsx bytes.

        One row per transaction (ref, date, amount, paid/pending) with a
        totals block below. `items` are VendorCardItem-shaped objects and
        `totals` carries total/paid_total/pending_total/count. Shared by all
        three card exports so their layout can never drift.
        """
        headers = ["Reference", "Date", "Amount", "Status"]

        def row_fn(it):
            return [
                getattr(it, "ref_no", ""),
                getattr(it, "transaction_date", None),
                getattr(it, "amount", Decimal("0.00")),
                str(getattr(it, "payment_state", "")).capitalize(),
            ]

        wb = self._build_workbook(
            sheet_name=(title[:31] or "Card"),
            headers=headers,
            records=items,
            row_fn=row_fn,
            money_cols=[3],
            date_cols=[2],
        )
        ws = wb.active

        # Title banner above the frozen header.
        ws.insert_rows(1)
        ws.merge_cells("A1:D1")
        banner = ws.cell(row=1, column=1, value=f"{vendor_name} — {title}")
        banner.font = Font(name="Calibri", bold=True, size=14, color="2C1A54")
        banner.alignment = Alignment(horizontal="left", vertical="center")
        ws.freeze_panes = "A3"

        # Totals block below the data.
        start = ws.max_row + 2
        summary_rows = [
            ("Total", totals.get("total")),
            ("Paid", totals.get("paid_total")),
            ("Pending", totals.get("pending_total")),
            ("Count", totals.get("count")),
        ]
        for offset, (label, value) in enumerate(summary_rows):
            r = start + offset
            lbl = ws.cell(row=r, column=1, value=label)
            lbl.font = Font(name="Calibri", bold=True, size=10)
            cell = ws.cell(row=r, column=3)
            if isinstance(value, Decimal):
                cell.value = float(value)
                cell.number_format = _MONEY_FORMAT
            else:
                cell.value = value
            cell.font = Font(name="Calibri", bold=True, size=10)
            cell.alignment = Alignment(horizontal="right")

        note = ws.cell(
            row=start + len(summary_rows) + 1,
            column=1,
            value=f"Currency: {currency}",
        )
        note.font = Font(name="Calibri", italic=True, size=9, color="817D7D")

        return self._to_bytes(wb)

    def export_customer_statement(self, statement) -> bytes:
        """Export a customer statement of accounts to .xlsx bytes."""

        def get(obj, key, default=None):
            if isinstance(obj, dict):
                return obj.get(key, default)
            return getattr(obj, key, default)

        customer = get(statement, "customer")
        summary = get(statement, "summary")
        transactions = get(statement, "transactions", [])

        customer_name = (
            get(customer, "company_name")
            or f"{get(customer, 'first_name', '')} {get(customer, 'last_name', '')}".strip()
            or get(customer, "email")
            or "Customer"
        )

        currency = get(customer, "currency", "")

        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"

        ws.merge_cells("A1:E1")
        title = ws.cell(
            row=1, column=1, value=f"{customer_name} — Statement of Accounts"
        )
        title.font = Font(name="Calibri", bold=True, size=14, color="2C1A54")
        title.alignment = Alignment(horizontal="left", vertical="center")

        ws.cell(row=2, column=1, value="Period Start")
        ws.cell(row=2, column=2, value=get(statement, "period_start"))
        ws.cell(row=3, column=1, value="Period End")
        ws.cell(row=3, column=2, value=get(statement, "period_end"))

        for row in (2, 3):
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True)
            ws.cell(row=row, column=2).number_format = _DATE_FORMAT

        summary_start = 5
        summary_rows = [
            ("Opening Balance", get(summary, "opening_balance")),
            ("Invoiced Amount", get(summary, "invoiced_amount")),
            ("Amount Paid", get(summary, "amount_paid")),
            ("Balance Due", get(summary, "balance_due")),
        ]

        for offset, (label, value) in enumerate(summary_rows):
            row = summary_start + offset
            ws.cell(row=row, column=1, value=label).font = Font(
                name="Calibri", bold=True
            )
            value_cell = ws.cell(row=row, column=2, value=float(value or 0))
            value_cell.number_format = _MONEY_FORMAT
            value_cell.alignment = Alignment(horizontal="right")

        header_row = summary_start + len(summary_rows) + 2
        headers = ["Date", "Details", "Amount", "Payment", "Balance"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        row_idx = header_row + 1

        for tx in transactions:
            ws.cell(row=row_idx, column=1, value=get(tx, "date"))
            ws.cell(row=row_idx, column=2, value=get(tx, "description", ""))
            ws.cell(row=row_idx, column=3, value=float(get(tx, "amount", 0) or 0))
            ws.cell(row=row_idx, column=4, value=float(get(tx, "payment", 0) or 0))
            ws.cell(row=row_idx, column=5, value=float(get(tx, "balance", 0) or 0))

            ws.cell(row=row_idx, column=1).number_format = _DATE_FORMAT

            for col_idx in (3, 4, 5):
                ws.cell(row=row_idx, column=col_idx).number_format = _MONEY_FORMAT
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    horizontal="right"
                )

            for col_idx in range(1, 6):
                ws.cell(row=row_idx, column=col_idx).font = _BODY_FONT
                ws.cell(row=row_idx, column=col_idx).border = _THIN_BORDER

            row_idx += 1

        row_idx += 1
        ws.cell(row=row_idx, column=4, value="Balance Due").font = Font(
            name="Calibri", bold=True
        )
        balance_cell = ws.cell(
            row=row_idx,
            column=5,
            value=float(get(summary, "balance_due", 0) or 0),
        )
        balance_cell.font = Font(name="Calibri", bold=True)
        balance_cell.number_format = _MONEY_FORMAT
        balance_cell.alignment = Alignment(horizontal="right")

        note = ws.cell(row=row_idx + 2, column=1, value=f"Currency: {currency}")
        note.font = Font(name="Calibri", italic=True, size=9, color="817D7D")

        widths = {
            "A": 16,
            "B": 45,
            "C": 16,
            "D": 16,
            "E": 16,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = f"A{header_row + 1}"

        return self._to_bytes(wb)

    # private implementation

    def _build_workbook(
        self,
        *,
        sheet_name: str,
        headers: list[str],
        records: list,
        row_fn,
        money_cols: list[int],
        date_cols: list[int],
    ) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # write header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        # write data rows
        for row_idx, record in enumerate(records, 2):
            values = row_fn(record)
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = _BODY_FONT
                cell.border = _THIN_BORDER

                if isinstance(value, Decimal):
                    cell.value = float(value)
                elif isinstance(value, (date, datetime)):
                    cell.value = value
                else:
                    cell.value = str(value) if value is not None else ""

                if col_idx in money_cols:
                    cell.number_format = _MONEY_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in date_cols:
                    cell.number_format = _DATE_FORMAT

        # auto-fit column widths
        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1])
            for row_idx in range(2, min(len(records) + 2, 52)):  # sample first 50 rows
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[letter].width = min(max_len + 3, 40)

        # freeze header row
        ws.freeze_panes = "A2"

        logger.info(
            "Built Excel sheet '%s' with %d records",
            sheet_name,
            len(records),
        )
        return wb

    def _add_line_items_sheet(
        self,
        wb: Workbook,
        records: list,
        doc_type: str,
    ) -> None:
        """Add a second sheet with line item details."""
        ws = wb.create_sheet("Line Items")

        headers = [
            f"{doc_type.title()} #",
            "Line #",
            "Item",
            "Description",
            "Qty",
            "Unit Price",
            "Tax Type",
            "Tax Amount",
            "Line Total",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN

        row_idx = 2
        for record in records:
            number_field = f"{doc_type}_number"
            doc_number = getattr(record, number_field, "")
            for item in getattr(record, "line_items", []):
                ws.cell(row=row_idx, column=1, value=doc_number).font = _BODY_FONT
                ws.cell(row=row_idx, column=2, value=item.line_number).font = _BODY_FONT
                ws.cell(row=row_idx, column=3, value=item.item_name).font = _BODY_FONT
                ws.cell(row=row_idx, column=4, value=item.description).font = _BODY_FONT

                qty_cell = ws.cell(row=row_idx, column=5, value=float(item.quantity))
                qty_cell.number_format = _MONEY_FORMAT
                qty_cell.font = _BODY_FONT

                price_cell = ws.cell(
                    row=row_idx, column=6, value=float(item.unit_price)
                )
                price_cell.number_format = _MONEY_FORMAT
                price_cell.font = _BODY_FONT

                ws.cell(
                    row=row_idx, column=7, value=str(item.tax_type)
                ).font = _BODY_FONT

                tax_cell = ws.cell(row=row_idx, column=8, value=float(item.tax_amount))
                tax_cell.number_format = _MONEY_FORMAT
                tax_cell.font = _BODY_FONT

                total_cell = ws.cell(
                    row=row_idx, column=9, value=float(item.line_total)
                )
                total_cell.number_format = _MONEY_FORMAT
                total_cell.font = _BODY_FONT

                row_idx += 1

        ws.freeze_panes = "A2"

    @staticmethod
    def _to_bytes(wb: Workbook) -> bytes:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()
