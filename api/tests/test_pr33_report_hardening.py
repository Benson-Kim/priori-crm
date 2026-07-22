"""Focused regression coverage for the final PR #33 report hardening."""

from datetime import date
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from starlette.requests import Request

from app.common.dependencies import get_reports_service
from app.common.excel import ExcelExporter
from app.common.exceptions import BadRequestException
from app.constants.enums import Currency, UserRole
from app.lib.config import settings
from app.modules.reports import audit as report_audit
from app.modules.reports.schemas import TaxByTypeRow
from app.modules.reports.service import ReportsService
from app.modules.statements.schemas import RangePreset, ResolvedPeriod


def _period() -> ResolvedPeriod:
    return ResolvedPeriod.resolve(
        RangePreset.CUSTOM,
        date(2026, 1, 1),
        date(2026, 1, 31),
    )


def _sales_row(index: int) -> SimpleNamespace:
    return SimpleNamespace(
        id=uuid4(),
        customer_name=f"Customer {index}",
        reference=f"REF-{index}",
        number=f"INV-{index}",
        date=date(2026, 1, 1),
        status="sent",
        currency="KES",
        subtotal=Decimal("100.00"),
        discount=Decimal("0.00"),
        net_revenue=Decimal("100.00"),
        amount=Decimal("116.00"),
        tax=Decimal("16.00"),
        balance_due=Decimal("116.00"),
    )


def test_export_size_accepts_limit_and_rejects_limit_plus_one(monkeypatch):
    monkeypatch.setattr(settings, "TAX_REPORT_EXPORT_MAX_ROWS", 2)

    ReportsService.validate_export_size(2, "Sales")
    with pytest.raises(BadRequestException):
        ReportsService.validate_export_size(3, "Sales")


def test_live_sales_export_guard_rejects_rows_beyond_limit(monkeypatch):
    monkeypatch.setattr(settings, "TAX_REPORT_EXPORT_MAX_ROWS", 2)
    service = ReportsService.__new__(ReportsService)
    service.repo = SimpleNamespace(
        iter_sales_ledger=lambda *args, **kwargs: iter(
            [_sales_row(1), _sales_row(2), _sales_row(3)]
        )
    )

    rows = service.export_sales_ledger(_period(), Currency.KES, batch_size=100)
    assert next(rows).number == "INV-1"
    assert next(rows).number == "INV-2"
    with pytest.raises(BadRequestException):
        next(rows)


def test_report_export_can_save_directly_to_file(tmp_path: Path):
    destination = tmp_path / "sales.xlsx"

    result = ExcelExporter().export_sales_report(
        [_sales_row(1)],
        "KES",
        destination=str(destination),
    )

    assert result is None
    assert destination.exists()
    workbook = load_workbook(destination, read_only=True, data_only=False)
    assert workbook["Sales"]["B2"].value == "REF-1"


def test_tax_breakdown_requires_and_preserves_taxable_value():
    row = TaxByTypeRow(
        tax_type="vat_0",
        tax_rate=Decimal("0"),
        taxable_value=Decimal("1250.00"),
        tax_amount=Decimal("0.00"),
        document_count=2,
    )

    assert row.taxable_value == Decimal("1250.00")


def test_tax_workbook_contains_taxable_value_columns(tmp_path: Path):
    destination = tmp_path / "vat.xlsx"
    row = TaxByTypeRow(
        tax_type="vat_16",
        tax_rate=Decimal("0.16"),
        taxable_value=Decimal("100.00"),
        tax_amount=Decimal("16.00"),
        document_count=1,
    )
    report = SimpleNamespace(
        metrics=SimpleNamespace(
            vat_collected=Decimal("16.00"),
            input_vat_estimate=Decimal("0.00"),
            net_vat_estimate=Decimal("16.00"),
        ),
        filing_warning="VAT reconciliation estimate only",
        completeness=SimpleNamespace(
            status="complete",
            excluded_document_count=0,
            excluded_currencies=[],
        ),
        sales_by_tax_type=[row],
        purchases_by_tax_type=[row],
    )

    result = ExcelExporter().export_tax_report(
        report,
        destination=str(destination),
    )

    assert result is None
    workbook = load_workbook(destination, read_only=True, data_only=False)
    assert workbook["Sales by Tax Type"]["B1"].value == "Taxable Value (KES)"
    assert workbook["Sales by Tax Type"]["B2"].value == 100
    assert workbook["Expense VAT by Type"]["B1"].value == "Taxable Value (KES)"


class _FakeDialect:
    name = "postgresql"


class _FakeBind:
    dialect = _FakeDialect()


class _FakeSession:
    def __init__(self) -> None:
        self.transaction_active = True
        self.commits = 0
        self.rollbacks = 0
        self.connection_options: list[dict] = []
        self.statements: list[str] = []

    def in_transaction(self) -> bool:
        return self.transaction_active

    def commit(self) -> None:
        self.commits += 1
        self.transaction_active = False

    def rollback(self) -> None:
        self.rollbacks += 1
        self.transaction_active = False

    def get_bind(self):
        return _FakeBind()

    def connection(self, *, execution_options: dict):
        self.connection_options.append(execution_options)
        self.transaction_active = True
        return self

    def execute(self, statement):
        self.statements.append(str(statement))
        return None


def _request(path: str = "/api/v1/reports/sales/export") -> Request:
    return Request(
        {
            "type": "http",
            "method": "GET",
            "path": path,
            "query_string": (
                b"range=custom&dateFrom=2026-01-01&dateTo=2026-01-31&currency=KES"
            ),
            "headers": [],
        }
    )


def test_report_service_starts_postgres_repeatable_read_snapshot():
    db = _FakeSession()
    service = get_reports_service(db, SimpleNamespace())

    assert service.db is db
    assert db.commits == 1
    assert db.connection_options == [{"isolation_level": "REPEATABLE READ"}]
    assert db.statements == ["SET TRANSACTION READ ONLY"]


def test_report_audit_success_uses_same_session(monkeypatch):
    db = _FakeSession()
    request = _request()
    request.state.report_export_metrics = {"exported_data_row_count": 7}
    actor = SimpleNamespace(id=uuid4(), role=UserRole.ADMIN)
    captured: dict = {}

    def append_audit(session, **kwargs):
        captured.update(kwargs)
        captured["session"] = session
        session.transaction_active = True

    monkeypatch.setattr(report_audit, "_append_audit", append_audit)
    dependency = report_audit.authorize_and_audit_report_access(request, db, actor)

    next(dependency)
    with pytest.raises(StopIteration):
        next(dependency)

    assert captured["session"] is db
    assert captured["payload"]["success"] is True
    assert captured["payload"]["export_metrics"] == {"exported_data_row_count": 7}
    assert db.commits == 2
    assert db.rollbacks == 0


def test_report_audit_failure_preserves_original_error(monkeypatch):
    db = _FakeSession()
    request = _request("/api/v1/reports/taxes")
    actor = SimpleNamespace(id=uuid4(), role=UserRole.MANAGER)
    captured: dict = {}

    def append_audit(session, **kwargs):
        captured.update(kwargs)
        captured["session"] = session
        session.transaction_active = True

    monkeypatch.setattr(report_audit, "_append_audit", append_audit)
    dependency = report_audit.authorize_and_audit_report_access(request, db, actor)

    next(dependency)
    original = RuntimeError("report failed")
    with pytest.raises(RuntimeError, match="report failed"):
        dependency.throw(original)

    assert captured["session"] is db
    assert captured["payload"]["success"] is False
    assert captured["payload"]["failure"] == "RuntimeError"
    assert db.rollbacks == 1
    assert db.commits == 1


def test_successful_report_fails_closed_and_removes_file_on_audit_error(
    monkeypatch, tmp_path: Path
):
    db = _FakeSession()
    request = _request()
    artifact = tmp_path / "report.xlsx"
    artifact.write_bytes(b"xlsx")
    request.state.report_export_path = str(artifact)
    actor = SimpleNamespace(id=uuid4(), role=UserRole.ADMIN)

    def fail_audit(*args, **kwargs):
        raise OSError("audit unavailable")

    monkeypatch.setattr(report_audit, "_append_audit", fail_audit)
    dependency = report_audit.authorize_and_audit_report_access(request, db, actor)

    next(dependency)
    with pytest.raises(OSError, match="audit unavailable"):
        next(dependency)

    assert not artifact.exists()
    assert db.rollbacks == 1
