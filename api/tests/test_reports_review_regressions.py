"""Regression coverage for PR #33 review findings."""

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import load_workbook

from app.common.excel import ExcelExporter
from app.constants.enums import (
    Currency,
    ExpenseStatus,
    InvoiceStatus,
    PurchaseOrderStatus,
    TaxType,
)
from app.modules.customers.models import Customer
from app.modules.expenses.models import Expense, ExpenseLineItem
from app.modules.invoices.models import Invoice, InvoiceLineItem
from app.modules.purchase_orders.models import PurchaseOrder
from app.modules.reports.service import ReportsService
from app.modules.statements.schemas import RangePreset, ResolvedPeriod
from app.modules.vendors.models import Vendor

PERIOD = ResolvedPeriod.resolve(
    RangePreset.CUSTOM,
    date(2026, 1, 1),
    date(2026, 1, 31),
)
AS_OF_DATE = date(2026, 1, 31)


def _customer(db, name: str = "Acme Ltd") -> Customer:
    customer = Customer(
        customer_type="business",
        company_name=name,
        first_name="Jane",
        last_name="Doe",
        email=f"{uuid4().hex}@example.test",
        phone="0700000000",
        currency=Currency.KES,
    )
    db.add(customer)
    db.flush()
    return customer


def _vendor(db, name: str = "Supplier Ltd") -> Vendor:
    vendor = Vendor(
        vendor_name=name,
        email=f"{uuid4().hex}@vendor.test",
        status="active",
        currency=Currency.KES,
    )
    db.add(vendor)
    db.flush()
    return vendor


def _invoice(
    db,
    customer: Customer,
    *,
    subtotal: Decimal,
    tax_total: Decimal,
    tax_type: TaxType,
    line_tax: Decimal,
    vat_enabled: bool = False,
    vat_rate: Decimal | None = None,
    transaction_date: date = date(2026, 1, 10),
) -> Invoice:
    total = subtotal + tax_total
    invoice = Invoice(
        invoice_number=f"INV-{uuid4().hex[:12]}",
        invoice_reference=f"IN-{uuid4().hex[:8]}",
        customer_id=customer.id,
        transaction_date=transaction_date,
        due_date=transaction_date + timedelta(days=30),
        status=InvoiceStatus.SENT,
        currency=Currency.KES,
        subtotal=subtotal,
        vat_enabled=vat_enabled,
        vat_rate=vat_rate,
        tax_total=tax_total,
        total_due=total,
        amount_paid=Decimal("0.00"),
        balance_due=total,
    )
    db.add(invoice)
    db.flush()
    db.add(
        InvoiceLineItem(
            invoice_id=invoice.id,
            line_number=1,
            item_name="Service",
            description="Service",
            quantity=Decimal("1"),
            unit_price=subtotal,
            line_total=subtotal,
            tax_type=tax_type,
            tax_amount=line_tax,
        )
    )
    db.flush()
    return invoice


def _expense(
    db,
    vendor: Vendor,
    *,
    subtotal: Decimal,
    tax_total: Decimal,
    status: ExpenseStatus,
    balance_due: Decimal,
    tax_type: TaxType = TaxType.NO_TAX,
    line_tax: Decimal = Decimal("0.00"),
) -> Expense:
    total = subtotal + tax_total
    expense = Expense(
        expense_number=f"EXP-{uuid4().hex[:12]}",
        expense_reference=f"EX-{uuid4().hex[:8]}",
        vendor_id=vendor.id,
        expense_date=date(2026, 1, 11),
        due_date=date(2026, 1, 25),
        status=status,
        currency=Currency.KES,
        subtotal=subtotal,
        tax_total=tax_total,
        total_due=total,
        amount_paid=total - balance_due,
        balance_due=balance_due,
    )
    db.add(expense)
    db.flush()
    db.add(
        ExpenseLineItem(
            expense_id=expense.id,
            line_number=1,
            item_name="Purchase",
            description="Purchase",
            quantity=Decimal("1"),
            unit_price=subtotal,
            line_total=subtotal,
            tax_type=tax_type,
            tax_amount=line_tax,
        )
    )
    db.flush()
    return expense


def _po(
    db,
    vendor: Vendor,
    *,
    subtotal: Decimal,
    tax_total: Decimal,
    status: PurchaseOrderStatus,
    balance_due: Decimal,
    vat_rate: Decimal | None,
) -> PurchaseOrder:
    total = subtotal + tax_total
    po = PurchaseOrder(
        po_number=f"PO-{uuid4().hex[:12]}",
        po_reference=f"PO-{uuid4().hex[:8]}",
        vendor_id=vendor.id,
        order_date=date(2026, 1, 12),
        status=status,
        currency=Currency.KES,
        subtotal=subtotal,
        vat_enabled=vat_rate is not None,
        vat_rate=vat_rate,
        tax_total=tax_total,
        total=total,
        amount_paid=total - balance_due,
        balance_due=balance_due,
    )
    db.add(po)
    db.flush()
    return po


def test_tax_breakdowns_reconcile_document_line_zero_and_custom_vat(db):
    customer = _customer(db)
    vendor = _vendor(db)

    _invoice(
        db,
        customer,
        subtotal=Decimal("100.00"),
        tax_total=Decimal("16.00"),
        tax_type=TaxType.VAT_16,
        line_tax=Decimal("16.00"),
    )
    _invoice(
        db,
        customer,
        subtotal=Decimal("200.00"),
        tax_total=Decimal("32.00"),
        tax_type=TaxType.NO_TAX,
        line_tax=Decimal("0.00"),
        vat_enabled=True,
        vat_rate=Decimal("0.1600"),
    )
    _invoice(
        db,
        customer,
        subtotal=Decimal("300.00"),
        tax_total=Decimal("0.00"),
        tax_type=TaxType.NO_TAX,
        line_tax=Decimal("0.00"),
        vat_enabled=True,
        vat_rate=Decimal("0.0000"),
    )
    _invoice(
        db,
        customer,
        subtotal=Decimal("100.00"),
        tax_total=Decimal("15.00"),
        tax_type=TaxType.NO_TAX,
        line_tax=Decimal("0.00"),
        vat_enabled=True,
        vat_rate=Decimal("0.1500"),
    )

    _expense(
        db,
        vendor,
        subtotal=Decimal("100.00"),
        tax_total=Decimal("16.00"),
        status=ExpenseStatus.PENDING,
        balance_due=Decimal("116.00"),
        tax_type=TaxType.VAT_16,
        line_tax=Decimal("16.00"),
    )
    _po(
        db,
        vendor,
        subtotal=Decimal("50.00"),
        tax_total=Decimal("0.00"),
        status=PurchaseOrderStatus.SENT,
        balance_due=Decimal("50.00"),
        vat_rate=Decimal("0.0000"),
    )
    _po(
        db,
        vendor,
        subtotal=Decimal("100.00"),
        tax_total=Decimal("15.00"),
        status=PurchaseOrderStatus.SENT,
        balance_due=Decimal("115.00"),
        vat_rate=Decimal("0.1500"),
    )
    db.commit()

    report = ReportsService(db).get_tax_report(PERIOD, Currency.KES)
    sales = {(r.tax_type, r.tax_rate): r for r in report.sales_by_tax_type}
    purchases = {(r.tax_type, r.tax_rate): r for r in report.purchases_by_tax_type}

    assert sales[("vat_16", Decimal("0.1600"))].tax_amount == Decimal("48.00")
    assert sales[("vat_16", Decimal("0.1600"))].document_count == 2
    assert sales[("vat_0", Decimal("0.0000"))].document_count == 1
    assert sales[("vat_0", Decimal("0.0000"))].tax_amount == Decimal("0.00")
    assert sales[("vat_custom", Decimal("0.1500"))].tax_amount == Decimal("15.00")
    assert (
        sum((r.tax_amount for r in report.sales_by_tax_type), Decimal("0"))
        == report.metrics.vat_collected
    )

    assert purchases[("vat_16", Decimal("0.1600"))].document_count == 1
    assert ("vat_0", Decimal("0.0000")) not in purchases
    assert ("vat_custom", Decimal("0.1500")) not in purchases
    assert (
        sum((r.tax_amount for r in report.purchases_by_tax_type), Decimal("0"))
        == report.metrics.input_vat_estimate
    )
    assert report.metrics.input_vat_estimate == Decimal("16.00")
    assert report.metrics.net_vat_estimate == Decimal("47.00")

    payload = ExcelExporter().export_tax_report(report, "KES")
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    sales_rows = list(workbook["Sales by Tax Type"].iter_rows(values_only=True))
    purchase_rows = list(workbook["Expense VAT by Type"].iter_rows(values_only=True))
    assert ("15% VAT (Custom)", 15, 1) in sales_rows
    assert ("0% VAT (Zero-rated)", 0, 1) in sales_rows
    assert ("16% VAT", 16, 1) in purchase_rows


def test_empty_report_contracts_expose_zero_values(db):
    service = ReportsService(db)

    tax = service.get_tax_report(PERIOD, Currency.KES)
    assert tax.metrics.vat_collected == Decimal("0.00")
    assert tax.metrics.input_vat_estimate == Decimal("0.00")
    assert tax.metrics.net_vat_estimate == Decimal("0.00")
    assert tax.report_label == "VAT reconciliation estimate"

    purchases = service.get_purchases_summary(PERIOD, Currency.KES)
    assert purchases.metrics.actual_spend == Decimal("0.00")
    assert purchases.metrics.po_commitments == Decimal("0.00")
    assert purchases.metrics.input_vat_estimate == Decimal("0.00")
    assert purchases.metrics.po_commitment_tax == Decimal("0.00")
    assert purchases.metrics.expense_count == 0
    assert purchases.metrics.po_commitment_count == 0
    assert purchases.metrics.outstanding_payables == Decimal("0.00")


def test_outstanding_payables_exclude_credits_and_zero_balances(db):
    vendor = _vendor(db)
    _expense(
        db,
        vendor,
        subtotal=Decimal("100.00"),
        tax_total=Decimal("0.00"),
        status=ExpenseStatus.PENDING,
        balance_due=Decimal("100.00"),
    )
    _expense(
        db,
        vendor,
        subtotal=Decimal("100.00"),
        tax_total=Decimal("0.00"),
        status=ExpenseStatus.PAID,
        balance_due=Decimal("-20.00"),
    )
    _po(
        db,
        vendor,
        subtotal=Decimal("50.00"),
        tax_total=Decimal("0.00"),
        status=PurchaseOrderStatus.SENT,
        balance_due=Decimal("50.00"),
        vat_rate=None,
    )
    _po(
        db,
        vendor,
        subtotal=Decimal("50.00"),
        tax_total=Decimal("0.00"),
        status=PurchaseOrderStatus.SENT,
        balance_due=Decimal("0.00"),
        vat_rate=None,
    )
    _po(
        db,
        vendor,
        subtotal=Decimal("50.00"),
        tax_total=Decimal("0.00"),
        status=PurchaseOrderStatus.PAID,
        balance_due=Decimal("-10.00"),
        vat_rate=None,
    )
    db.commit()

    service = ReportsService(db)
    summary = service.get_purchases_summary(PERIOD, Currency.KES)
    aging = service.get_aged_payables(Currency.KES.value, AS_OF_DATE)
    assert summary.metrics.actual_spend == Decimal("200.00")
    assert summary.metrics.po_commitments == Decimal("150.00")
    assert summary.metrics.outstanding_payables == Decimal("100.00")
    assert aging.totals.total == Decimal("100.00")
    assert aging.vendors[0].total == Decimal("100.00")


def test_report_export_consumes_every_row_and_neutralizes_formula_text(db):
    class FakeRepository:
        batch_size: int | None = None

        def iter_sales_ledger(self, *args, batch_size: int, **kwargs):
            self.batch_size = batch_size
            for index in range(125):
                yield SimpleNamespace(
                    id=uuid4(),
                    customer_name="=2+2" if index == 0 else f"Customer {index}",
                    reference=f"IN-{index:04d}",
                    number=f"INV-{index:04d}",
                    date=date(2026, 1, 1) + timedelta(days=index % 31),
                    status=InvoiceStatus.SENT,
                    currency=Currency.KES,
                    subtotal=Decimal("100.00"),
                    discount=Decimal("0.00"),
                    net_revenue=Decimal("100.00"),
                    amount=Decimal("116.00"),
                    tax=Decimal("16.00"),
                    balance_due=Decimal("116.00"),
                )

    service = ReportsService(db)
    fake = FakeRepository()
    service.repo = fake

    rows = service.export_sales_ledger(PERIOD, Currency.KES, batch_size=17)
    payload = ExcelExporter().export_sales_report(rows, "KES")

    assert fake.batch_size == 17
    workbook = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    worksheet = workbook["Sales"]
    exported_rows = list(worksheet.iter_rows(values_only=True))
    # Header plus every row; greater than the API maximum interactive page size.
    assert len(exported_rows) == 126
    assert exported_rows[1][3] == "'=2+2"


def test_duplicate_customer_and_vendor_names_remain_separate_in_aging(db):
    customer_one = _customer(db, "Same Name Ltd")
    customer_two = _customer(db, "Same Name Ltd")
    _invoice(
        db,
        customer_one,
        subtotal=Decimal("100.00"),
        tax_total=Decimal("0.00"),
        tax_type=TaxType.NO_TAX,
        line_tax=Decimal("0.00"),
    )
    _invoice(
        db,
        customer_two,
        subtotal=Decimal("200.00"),
        tax_total=Decimal("0.00"),
        tax_type=TaxType.NO_TAX,
        line_tax=Decimal("0.00"),
    )

    vendor_one = _vendor(db, "Same Supplier Ltd")
    vendor_two = _vendor(db, "Same Supplier Ltd")
    _expense(
        db,
        vendor_one,
        subtotal=Decimal("100.00"),
        tax_total=Decimal("0.00"),
        status=ExpenseStatus.PENDING,
        balance_due=Decimal("100.00"),
    )
    _expense(
        db,
        vendor_two,
        subtotal=Decimal("50.00"),
        tax_total=Decimal("0.00"),
        status=ExpenseStatus.PENDING,
        balance_due=Decimal("50.00"),
    )
    db.commit()

    service = ReportsService(db)
    sales_summary = service.get_sales_summary(PERIOD, Currency.KES)
    purchase_summary = service.get_purchases_summary(PERIOD, Currency.KES)
    receivables = service.get_aged_receivables(Currency.KES.value, AS_OF_DATE)
    payables = service.get_aged_payables(Currency.KES.value, AS_OF_DATE)

    assert {row.customer_id for row in sales_summary.revenue_by_customer} == {
        customer_one.id,
        customer_two.id,
    }
    assert {row.vendor_id for row in purchase_summary.spend_by_vendor} == {
        vendor_one.id,
        vendor_two.id,
    }
    assert len(receivables.customers) == 2
    assert {row.customer_id for row in receivables.customers} == {
        customer_one.id,
        customer_two.id,
    }
    assert len(payables.vendors) == 2
    assert {row.vendor_id for row in payables.vendors} == {vendor_one.id, vendor_two.id}


def test_historical_vat_8_invoice_remains_reportable(db):
    customer = _customer(db, "Historical Customer")
    historical_period = ResolvedPeriod.resolve(
        RangePreset.CUSTOM,
        date(2023, 6, 1),
        date(2023, 6, 30),
    )
    _invoice(
        db,
        customer,
        subtotal=Decimal("100.00"),
        tax_total=Decimal("8.00"),
        tax_type=TaxType.VAT_8,
        line_tax=Decimal("8.00"),
        transaction_date=date(2023, 6, 30),
    )
    db.commit()

    report = ReportsService(db).get_tax_report(historical_period, Currency.KES)
    rows = {(row.tax_type, row.tax_rate): row for row in report.sales_by_tax_type}
    historical = rows[("vat_8", Decimal("0.0800"))]
    assert historical.tax_amount == Decimal("8.00")
    assert historical.document_count == 1
